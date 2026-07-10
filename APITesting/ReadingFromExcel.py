from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

# Load Excel
workbook = load_workbook("APITestData.xlsx")
sheet = workbook.active

with sync_playwright() as p:

    request = p.request.new_context()

    # Loop through all rows except header
    for row in range(2, sheet.max_row + 1):

        payload = {
            "userId": sheet.cell(row, 1).value,
            "title": sheet.cell(row, 2).value,
            "body": sheet.cell(row, 3).value
        }

        response = request.post(
            "https://jsonplaceholder.typicode.com/posts",
            data=payload
        )

        response_data = response.json()

        # Write Status Code
        sheet.cell(row, 4).value = response.status

        # Write Response ID
        sheet.cell(row, 5).value = response_data["id"]

# Save the workbook
workbook.save("APITestData.xlsx")
print("Excel Updated Successfully")